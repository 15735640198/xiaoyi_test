#!/usr/bin/env python3
"""生成 HarmonyOS 设置 API 汇总 Excel"""
import sys
sys.stdout.reconfigure(encoding='utf-8', errors='replace')
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "HarmonyOS 设置 API 汇总"

# 数据: (类别, [(设置项, API), ...])
categories = [
    ("网络与连接", [
        ("WLAN", "query_wlan()\nset_wlan('on'/'off')\nconnect_wlan(ssid, password)"),
        ("WLAN下自动下载", "query_wlan_auto_download()\nset_wlan_auto_download('on'/'off')"),
        ("蓝牙开关", "query_bluetooth()\nset_bluetooth('on'/'off')"),
        ("蓝牙设备", "query_bluetooth_device(name)\nconnect_bluetooth(name)\ndisconnect_bluetooth(name)"),
        ("星闪", "query_nearlink()\nset_nearlink('on'/'off')"),
        ("飞行模式", "query_flight_mode()\nset_flight_mode('on'/'off')"),
        ("个人热点", "query_personal_hotspot()\nset_personal_hotspot('on'/'off')"),
        ("热点配置", "query_hotspot_config()\nset_hotspot_name(name)\nset_hotspot_password(pwd)"),
        ("已连接设备", "query_hotspot_connected_devices()"),
        ("AP 频段", "query_hotspot_ap_band()"),
        ("USB 共享网络", "query_usb_tethering()\nset_usb_tethering('on'/'off')"),
        ("网络加速", "query_network_acceleration()\nset_network_acceleration('on'/'off')"),
        ("NFC", "query_nfc()\nset_nfc('on'/'off')"),
        ("默认付款应用", "query_default_payment_app()\nset_default_payment_app(app_name)"),
    ]),
    ("移动网络与SIM卡", [
        ("默认数据卡", "query_default_data_card()\nset_default_data_card('1'/'2')"),
        ("SIM卡状态", "query_sim_status()\nquery_sim_carrier()"),
        ("SIM卡使用状态", "query_sim_enabled(card)\nset_sim_enabled(card, 'on'/'off')"),
        ("应用联网", "query_app_network_access(app_name)\nset_app_network_access(app_name, type, 'on'/'off')"),
        ("应用流量使用量", "query_app_data_usage(app_name, period='30d'/'24h')"),
    ]),
    ("显示与亮度", [
        ("屏幕亮度", "query_brightness() →百分比"),
        ("自动调节亮度", "query_auto_brightness()\nset_auto_brightness('on'/'off')"),
        ("字体大小", "query_font_size()\nset_font_size('小'/'标准'/'大'/'超大')"),
        ("字体粗细", "query_font_weight()\nset_font_weight('最细'/'标准'/'最粗')"),
        ("电子书模式", "query_ebook_mode()"),
    ]),
    ("声音", [
        ("朗读速度", "query_speech_rate()\nset_speech_rate(value)"),
        ("来电铃声", "query_ringtone()\nset_ringtone_default()"),
        ("信息铃声", "query_message_ringtone()"),
        ("通知铃声", "query_notification_ringtone()"),
        ("响铃时振动", "query_ring_vibration()\nset_ring_vibration('on'/'off')"),
    ]),
    ("电池", [
        ("省电模式", "query_power_saving()\nset_power_saving('on'/'off')"),
    ]),
    ("通知与免打扰", [
        ("勿扰模式", "query_dnd()\nset_dnd('on'/'off')"),
    ]),
    ("辅助功能", [
        ("放大手势", "query_zoom_gesture()\nset_zoom_gesture('on'/'off')"),
        ("颜色反转", "query_color_inversion()\nset_color_inversion('on'/'off')"),
    ]),
    ("系统", [
        ("系统导航模式", "query_navigation_mode()"),
        ("系统语言", "query_system_language()\nadd_system_language(lang)"),
        ("默认输入法", "query_default_input_method()"),
        ("自动时区", "query_auto_timezone()\nset_auto_timezone('on'/'off')"),
        ("系统时区", "query_timezone()\nset_timezone(timezone_name)"),
        ("时间制式", "query_time_format()\nset_time_format('24'/'12')"),
        ("系统日期", "query_date()"),
        ("系统时间", "query_time()\nset_time(hour, minute)"),
        ("存储空间", "query_storage()"),
        ("开发者模式", "query_developer_mode()\nset_developer_mode('on'/'off')"),
        ("USB调试", "query_usb_debug()\nset_usb_debug('on'/'off')"),
    ]),
    ("安全", [
        ("锁屏方式", "query_lock_screen_method()\n(安全验证，不可自动化)"),
        ("锁屏密码", "set_lock_screen_password(password)"),
        ("指纹录入状态", "query_fingerprint()\n(录入需物理传感器)"),
        ("隐私空间", "query_privacy_space()\nset_privacy_space(main_pwd, space_pwd)"),
    ]),
]

# 按类别内 API 数量降序排序
categories.sort(key=lambda x: len(x[1]), reverse=True)

# 样式
header_font = Font(name='微软雅黑', bold=True, size=12, color='FFFFFF')
header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
data_font = Font(name='微软雅黑', size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top', horizontal='left')
center_align = Alignment(wrap_text=True, vertical='center', horizontal='center')

# 写入: 每个类别占2列 (设置项 | API)
col = 1
max_rows = max(len(items) for _, items in categories)

# 第1行: 类别名 (合并2列)
for cat_name, items in categories:
    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 1)
    cell = ws.cell(row=1, column=col, value=cat_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    for c in range(col, col + 2):
        ws.cell(row=1, column=c).fill = header_fill
        ws.cell(row=1, column=c).border = thin_border
    col += 2

# 第2行: 子标题
col = 1
for _ in categories:
    for i, title in enumerate(["设置项", "API"]):
        cell = ws.cell(row=2, column=col + i, value=title)
        cell.font = Font(name='微软雅黑', bold=True, size=10)
        cell.fill = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    col += 2

# 数据行
for row_idx in range(max_rows):
    col = 1
    for cat_name, items in categories:
        if row_idx < len(items):
            name, api = items[row_idx]
            for i, val in enumerate([name, api]):
                cell = ws.cell(row=row_idx + 3, column=col + i, value=val)
                cell.font = data_font
                cell.alignment = wrap_align
                cell.border = thin_border
        else:
            for i in range(2):
                cell = ws.cell(row=row_idx + 3, column=col + i, value="")
                cell.border = thin_border
        col += 2

# 列宽
for c in range(1, len(categories) * 2 + 1):
    ws.column_dimensions[get_column_letter(c)].width = 30

# 行高
ws.row_dimensions[1].height = 30
ws.row_dimensions[2].height = 22
for r in range(3, max_rows + 3):
    ws.row_dimensions[r].height = 50

# 冻结前两行
ws.freeze_panes = 'A3'

output = r"D:\lzs\xiaoyi\github_repo\xiaoyi_test\HarmonyOS_API汇总.xlsx"
wb.save(output)
print(f"已保存: {output}")
